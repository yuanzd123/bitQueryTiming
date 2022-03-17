#!/usr/bin/python
# -*- coding: utf-8 -*-
import requests
import pandas as pd
import openpyxl
from openpyxl.utils.cell import get_column_letter

def run_query(query):  # A simple function to use requests.post to make the API call.
    headers = {'X-API-KEY': 'BQYPV3D5XIZ0UeT8usav7f1LtbQwjEYW'}
    request = requests.post('https://graphql.bitquery.io/',
                            json={'query': query}, headers=headers)
    if request.status_code == 200:
        return request.json()
    else:
        raise Exception('Query failed and return code is {}.      {}'.format(request.status_code,
                        query))


# The GraphQL query

query = """
{
  ethereum(network: bsc) {
    dexTrades(
      options:{limitBy:{each:"date.dt",limit:2},desc:"Total_USD"}
      quoteCurrency:{notIn:[
        "0xbb4CdB9CBd36B01bD1cBaEBF2De08d9173bc095c",
        "0x23396cf899ca06c4472205fc903bdb4de249d6fc",
        "0x55d398326f99059ff775485246999027b3197955", 
        "0xe9e7cea3dedca5984780bafc599bd69add087d56"
      ]}
      date: {since:"2022-03-15"}
      
    ) {
      exchange{name
      address{address}}
      baseCurrency {
        base_sc:address
        base: symbol
      }
      quoteCurrency{
        quote_sc:address
        quote:symbol
      }
      baseAmount
      quoteAmount
      date{dt:date}
       Total_USD: tradeAmount(calculate: sum, in: USD)
      Tx_count:count(uniq:txs)
      smartContract{address{sc:address}}
    
    }
  }
}
"""
# result = run_query(query)  # Execute the query
result = {'data': {'ethereum': {'dexTrades': [{'exchange': {'name': None, 'address': {'address': '0x137f34df5bcdb30f5e858fc77cb7ab60f8f7a09a'}}, 'baseCurrency': {'base_sc': '0x6b23c89196deb721e6fd9726e6c76e4810a464bc', 'base': 'XWG'}, 'quoteCurrency': {'quote_sc': '0x181801f00df1bd997d38dd579dbd44bf9b5a6d2d', 'quote': 'YOU'}, 'baseAmount': 2.6579755755937014e+19, 'quoteAmount': 97.461156, 'date': {'dt': '2022-03-16'}, 'Total_USD': 1.2516000397597678e+18, 'Tx_count': 1, 'smartContract': {'address': {'sc': '0x903426bcd04661ae7a32e451e0ac7f372d3bb7b6'}}}, {'exchange': {'name': None, 'address': {'address': '0x137f34df5bcdb30f5e858fc77cb7ab60f8f7a09a'}}, 'baseCurrency': {'base_sc': '0x181801f00df1bd997d38dd579dbd44bf9b5a6d2d', 'base': 'YOU'}, 'quoteCurrency': {'quote_sc': '0x6b23c89196deb721e6fd9726e6c76e4810a464bc', 'quote': 'XWG'}, 'baseAmount': 97.461156, 'quoteAmount': 2.6579755755937014e+19, 'date': {'dt': '2022-03-16'}, 'Total_USD': 1.2516000397597678e+18, 'Tx_count': 1, 'smartContract': {'address': {'sc': '0x903426bcd04661ae7a32e451e0ac7f372d3bb7b6'}}}, {'exchange': {'name': 'Pancake v2', 'address': {'address': '0xca143ce32fe78f1f7019d7d551a6402fc5350c73'}}, 'baseCurrency': {'base_sc': '0x23396cf899ca06c4472205fc903bdb4de249d6fc', 'base': 'UST'}, 'quoteCurrency': {'quote_sc': '0x6b23c89196deb721e6fd9726e6c76e4810a464bc', 'quote': 'XWG'}, 'baseAmount': 0.004777912937316232, 'quoteAmount': 7.652213677514035e+16, 'date': {'dt': '2022-03-15'}, 'Total_USD': 4042161017155593.0, 'Tx_count': 1, 'smartContract': {'address': {'sc': '0x06bb21bbadf4d650532046b92398da6462b8e5eb'}}}, {'exchange': {'name': None, 'address': {'address': '0x9a272d734c5a0d7d84e0a892e891a553e8066dce'}}, 'baseCurrency': {'base_sc': '0x55d398326f99059ff775485246999027b3197955', 'base': 'USDT'}, 'quoteCurrency': {'quote_sc': '0xc9882def23bc42d53895b8361d0b1edc7570bc6a', 'quote': 'FIST'}, 'baseAmount': 69236080.70221113, 'quoteAmount': 35662656.379215, 'date': {'dt': '2022-03-15'}, 'Total_USD': 69343316.64352229, 'Tx_count': 22714, 'smartContract': {'address': {'sc': '0xb4ec801aed8c92f2e69589518aaa127afb37d8c9'}}}, {'exchange': {'name': 'Pancake v2', 'address': {'address': '0xca143ce32fe78f1f7019d7d551a6402fc5350c73'}}, 'baseCurrency': {'base_sc': '0x55d398326f99059ff775485246999027b3197955', 'base': 'USDT'}, 'quoteCurrency': {'quote_sc': '0x26619fa1d4c957c58096bbbeca6588dcfb12e109', 'quote': 'TIME'}, 'baseAmount': 4684454.705493295, 'quoteAmount': 22211219.082697917, 'date': {'dt': '2022-03-17'}, 'Total_USD': 4691573.218440662, 'Tx_count': 3493, 'smartContract': {'address': {'sc': '0xad1fedfb04377c4b849cef6ef9627bca41955fa0'}}}, {'exchange': {'name': None, 'address': {'address': '0x9a272d734c5a0d7d84e0a892e891a553e8066dce'}}, 'baseCurrency': {'base_sc': '0x55d398326f99059ff775485246999027b3197955', 'base': 'USDT'}, 'quoteCurrency': {'quote_sc': '0xc9882def23bc42d53895b8361d0b1edc7570bc6a', 'quote': 'FIST'}, 'baseAmount': 2748486.561303439, 'quoteAmount': 1454176.817368, 'date': {'dt': '2022-03-17'}, 'Total_USD': 2752394.424677198, 'Tx_count': 2035, 'smartContract': {'address': {'sc': '0xb4ec801aed8c92f2e69589518aaa127afb37d8c9'}}}]}}}

print (type(result))
result = result.get('data').get('ethereum').get('dexTrades')

# for query in result:
#   for item in query.items():
#     print(item)
#     print(type(item))

df = pd.DataFrame.from_dict(result)
# print (df)
writer = pd.ExcelWriter('E:\Helix\dexTrades.xlsx', engine='xlsxwriter')
df.to_excel(writer, sheet_name='Sheet1', startrow=1, header=0,index=False)
workbook  = writer.book
worksheet = writer.sheets['Sheet1']
 

# Add a header format.
header_format = workbook.add_format({
    'bold': True, # 字体加粗
    'text_wrap': True, # 是否自动换行
    'valign': 'top',  #垂直对齐方式
    'align': 'right', # 水平对齐方式
    'fg_color': '#D7E4BC', # 单元格背景颜色
    'border': 2}) # 单元格边框宽度
 
 
yellow = workbook.add_format({'fg_color': '#FFEE99'})
red=workbook.add_format({'fg_color': '#2dB054'})
 
# Write the column headers with the defined format.
for col_num, value in enumerate(df.columns.values):
    if col_num%2==0:
        worksheet.write(0, col_num, value, header_format)
    else:
        worksheet.write(0, col_num, value, yellow)
 
# # Write the row with the defined format.
# for index, value in df.iterrows():
#     print(index," -- > ",value.values)
#     if index % 2 == 0:
#         worksheet.write(index+1, 0, value[0], header_format)
#     else:
#         worksheet.write(index+1, 0, value[0], yellow)
 
worksheet.set_column("A:C", 16)
format2 = workbook.add_format({'bold':  True, 'align': 'vcenter', 'valign': 'top', 'text_wrap': True})
worksheet.set_row(0, cell_format=format2)

amountOfRows = worksheet.max_row
amountOfColumns = worksheet.max_column

for i in range(amountOfColumns):
    for k in range(amountOfRows):
        cell = str(worksheet[get_column_letter(i+1)+str(k+1)].value)
        if( str(cell[0]) == "'address': {'address': '"):
            newCell = "address"+cell[1:]
            worksheet[get_column_letter(i+1)+str(k+1)]=newCell

writer.save()
print("gg")